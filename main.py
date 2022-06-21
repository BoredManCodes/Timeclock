import datetime
import tkinter as tk
import openpyxl

# read from the employee list
wb = openpyxl.load_workbook('timeclock_files/data.xlsx')
employee_data = wb.worksheets[0]
# create the window
window = tk.Tk()
window.geometry("600x600")
window.title("Employee TimeClock")
# create a title on screen
title = tk.Label(text="Scan your tag to clock in/out")
title.pack()
# create a text input for their employee ID
employee_id_input = tk.Entry(width=50)
employee_id_input.pack()
# we focus on it, so it's ready to go on start
employee_id_input.focus()
# create a label for displaying our last clock in
last_clock_in = tk.Label(text="yes")
last_clock_in.pack()
# create label for current day time
date_time = tk.Label(text="")
date_time.pack()


# when the card is read it automatically presses enter after, we hijack that here to run the rest of our code
def process_clock(event):
    # only check the second column which contains our IDs, and ignore the headers
    for row in employee_data.iter_rows(min_col=2, max_col=2, min_row=2):
        # iterating over rows is a shit idea, but we don't have that many employees so it *should* be fine
        for cell in row:
            # take the value of cells, convert to an integer, then convert it to a string,
            # then remove the ".0" from the end. If the output matches then we have found the user from the ID
            print(employee_id_input.get())
            if int(cell.value).__str__().split(".")[0] == employee_id_input.get():
                print(cell.value)
                ts = str(datetime.datetime.now()).split(".")[0]
                last_clock_in.configure(
                    text=f"Clocked in {employee_data.cell(cell.row, 1).value} at {str(ts).split('.')[0]}")
            else:
                print(cell.value)
                last_clock_in.configure(text="Couldn't find your ID in the database. Please see reception.")
    # clear the input for the next person
    employee_id_input.delete(0, tk.END)


def update():
    ts = str(datetime.datetime.now()).split(".")[0]
    date_time.configure(text=ts)
    window.after(1000, update)

# hook into the "return" key to run the above function
window.bind('<Return>', process_clock)

update()
# display the window and loop until closed
window.mainloop()
